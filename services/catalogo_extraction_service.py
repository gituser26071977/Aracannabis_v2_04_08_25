"""
Serviço de Extração de Catálogo por IA — Squad D (Integração SGA)
Adaptado do InventoryIntelligentService do SGA para o SIAP.

Usa o gateway LLM existente (services.ai_agents) — NÃO recria LLMClient.
Suporta: PDF, PNG, JPG, JPEG, XLSX.
"""

import os
import io
import base64
import json
import logging
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional

from models import db
from models_extra import CatalogoImportLog

logger = logging.getLogger(__name__)

# Prompt de sistema genérico para extração de produtos de catálogo médico/clínico
CATALOG_SYSTEM_PROMPT = """Você é um especialista em gestão de estoque de clínicas médicas.
Extraia uma lista de produtos deste documento/catálogo.
Para cada produto, identifique:
- name: Nome completo do produto (ex: Óleo CBD Full Spectrum 3000mg, Paracetamol 500mg, etc.)
- category: Categoria do produto (ex: óleo, flor, pomada, gummy, pet, vaporizador, medicamento, insumo, equipamento, suplemento, outro)
- description: Breve descrição ou concentração do produto
- unit: Unidade de medida (ex: ml, g, mg, un, comprimido, cápsula)
- concentration: Concentração principal, se indicada (ex: "3000mg", "500mg/ml", "10%")
- manufacturer: Fabricante ou marca, se indicada
- barcode: Código de barras, se presente

Retorne APENAS um JSON válido no formato:
[
  {
    "name": "...",
    "category": "...",
    "description": "...",
    "unit": "...",
    "concentration": "...",
    "manufacturer": "...",
    "barcode": "..."
  },
  ...
]

Se não encontrar algum campo, use string vazia "". Não inclua explicações fora do JSON."""


class CatalogoExtractionService:
    """Serviço de extração inteligente de produtos de catálogos via IA."""

    def __init__(self):
        from services.ai_agents import ai_manager
        self.ai_manager = ai_manager
        self.daily_limit = 10

    # REDACTED
    # Rate limiting
    # REDACTED
    def _check_rate_limit(self, user_id: int) -> bool:
        """Verifica se o usuário ainda pode fazer extrações hoje (trial)."""
        today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
        today_end = today_start + timedelta(days=1)
        count = CatalogoImportLog.query.filter(
            CatalogoImportLog.user_id == user_id,
            CatalogoImportLog.created_at >= today_start,
            CatalogoImportLog.created_at < today_end
        ).count()
        return count < self.daily_limit

    def _log_extraction(self, user_id: int, filename: Optional[str],
                        detected_count: int = 0, imported_count: int = 0,
                        errors: Optional[List[str]] = None) -> CatalogoImportLog:
        """Registra uma entrada no log de importação."""
        log = CatalogoImportLog(
            user_id=user_id,
            filename=filename,
            detected_count=detected_count,
            imported_count=imported_count,
            errors=errors or []
        )
        db.session.add(log)
        db.session.commit()
        return log

    # REDACTED
    # Processamento de arquivos
    # REDACTED
    def _extract_text_from_pdf(self, file_bytes: bytes) -> str:
        """Extrai texto de um PDF usando pypdf."""
        try:
            from pypdf import PdfReader
            reader = PdfReader(io.BytesIO(file_bytes))
            parts = []
            for page in reader.pages:
                text = page.extract_text()
                if text:
                    parts.append(text)
            return "\n\n".join(parts)
        except Exception as e:
            logger.warning(f"Falha ao extrair texto do PDF: {e}")
            return ""

    def _extract_text_from_xlsx(self, file_bytes: bytes) -> str:
        """Extrai dados de uma planilha XLSX como texto/markdown."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            parts = []
            for sheet in wb.worksheets:
                parts.append(f"--- Sheet: {sheet.title} ---")
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip():
                        rows.append(row_text)
                parts.append("\n".join(rows))
            return "\n\n".join(parts)
        except Exception as e:
            logger.warning(f"Falha ao extrair texto do XLSX: {e}")
            return ""

    def _file_to_base64(self, file_bytes: bytes) -> str:
        """Converte bytes para base64 puro (sem prefixo data URI)."""
        return base64.b64encode(file_bytes).decode("utf-8")

    # REDACTED
    # IA
    # REDACTED
    def _call_ia_for_text(self, text_content: str) -> List[Dict[str, Any]]:
        """Envia texto extraído para o LLM e retorna lista de produtos."""
        prompt = (
            "Extraia todos os produtos encontrados no texto do catálogo a seguir.\n\n"
            f"{text_content[:12000]}"  # limite generoso de contexto
        )
        messages = [
            {"role": "system", "content": CATALOG_SYSTEM_PROMPT},
            {"role": "user", "content": prompt},
        ]
        response = self.ai_manager.chat_completion(
            messages=messages,
            temperature=0.1,
            max_tokens=4000,
        )
        return self._parse_json_response(response.get("content", ""))

    def _call_ia_for_image(self, image_b64: str, mime_type: str) -> List[Dict[str, Any]]:
        """Envia imagem para o LLM de visão e retorna lista de produtos."""
        # Construir data URI correto
        if not image_b64.startswith("data:"):
            image_b64 = f"data:{mime_type};base64,{image_b64}"

        response = self.ai_manager.vision_completion(
            prompt=CATALOG_SYSTEM_PROMPT + "\n\nExtraia todos os produtos visíveis nesta imagem de catálogo.",
            image_data=image_b64,
            temperature=0.1,
            max_tokens=4000,
        )
        return self._parse_json_response(response.get("content", ""))

    def _parse_json_response(self, text: str) -> List[Dict[str, Any]]:
        """Limpa e faz parse do JSON retornado pela IA."""
        if not text:
            return []

        clean = text.strip()
        if clean.startswith("```json"):
            clean = clean[7:]
        if clean.startswith("```"):
            clean = clean[3:]
        if clean.endswith("```"):
            clean = clean[:-3]
        clean = clean.strip()

        try:
            parsed = json.loads(clean)
            if isinstance(parsed, list):
                return parsed
            if isinstance(parsed, dict):
                # Alguns LLMs retornam { "products": [...] }
                for key in ("products", "produtos", "items", "result", "data"):
                    if key in parsed and isinstance(parsed[key], list):
                        return parsed[key]
                return [parsed]
        except Exception as e:
            logger.error(f"Erro ao parsear JSON da IA: {e}. Texto: {clean[:300]}...")
        return []

    # REDACTED
    # Conversão para modelo SIAP
    # REDACTED
    @staticmethod
    def _map_to_siap(products: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """Mapeia campos extraídos para o formato do modelo Produto do SIAP."""
        mapped = []
        for p in products:
            mapped.append({
                "nome": p.get("name", "").strip(),
                "categoria": p.get("category", "").strip(),
                "descricao": p.get("description", "").strip(),
                "unidade": p.get("unit", "").strip(),
                "concentracao": p.get("concentration", "").strip(),
                "fabricante": p.get("manufacturer", "").strip(),
                "codigo_barras": p.get("barcode", "").strip(),
                "tipo": _infer_tipo(p.get("category", "")),
            })
        return mapped

    # REDACTED
    # API pública
    # REDACTED
    def extract_from_file(self, file_bytes: bytes, filename: str,
                          mime_type: str, user_id: int) -> Dict[str, Any]:
        """
        Fluxo principal:
        1. Verifica rate limit
        2. Extrai conteúdo do arquivo
        3. Chama IA
        4. Mapeia para modelo SIAP
        5. Registra log
        6. Retorna JSON
        """
        if not self._check_rate_limit(user_id):
            return {
                "error": "Limite diário atingido",
                "message": f"Você atingiu o limite de {self.daily_limit} extrações por dia.",
                "detected_products": [],
                "count": 0,
            }

        ext = os.path.splitext(filename.lower())[1]
        products = []
        error_msg = None

        try:
            if mime_type.startswith("image/") or ext in (".png", ".jpg", ".jpeg"):
                b64 = self._file_to_base64(file_bytes)
                products = self._call_ia_for_image(b64, mime_type or "image/jpeg")
            elif mime_type == "application/pdf" or ext == ".pdf":
                text = self._extract_text_from_pdf(file_bytes)
                if not text.strip():
                    return {
                        "error": "Não foi possível extrair texto do PDF",
                        "message": "O PDF pode estar protegido ou conter apenas imagens. Tente uma imagem do catálogo.",
                        "detected_products": [],
                        "count": 0,
                    }
                products = self._call_ia_for_text(text)
            elif mime_type in (
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            ) or ext in (".xlsx", ".xls"):
                text = self._extract_text_from_xlsx(file_bytes)
                if not text.strip():
                    return {
                        "error": "Não foi possível extrair dados da planilha",
                        "detected_products": [],
                        "count": 0,
                    }
                products = self._call_ia_for_text(text)
            else:
                return {
                    "error": "Formato não suportado",
                    "message": "Use PDF, PNG, JPG ou XLSX.",
                    "detected_products": [],
                    "count": 0,
                }
        except Exception as e:
            logger.exception("Erro na extração de catálogo por IA")
            error_msg = str(e)
            products = []

        mapped = self._map_to_siap(products)

        # Filtra produtos sem nome
        mapped = [p for p in mapped if p.get("nome")]

        # Registra log
        log_errors = [error_msg] if error_msg else None
        self._log_extraction(
            user_id=user_id,
            filename=filename,
            detected_count=len(mapped),
            imported_count=0,
            errors=log_errors,
        )

        if error_msg:
            return {
                "error": "Erro durante extração",
                "message": error_msg,
                "detected_products": mapped,
                "count": len(mapped),
            }

        return {
            "detected_products": mapped,
            "count": len(mapped),
        }

    def import_products(self, products: List[Dict[str, Any]], user_id: int,
                        filename: Optional[str] = None) -> Dict[str, Any]:
        """
        Salva produtos selecionados no banco como Produtos ativos.
        Retorna contagem de sucessos/erros.
        """
        from models import Produto

        imported = 0
        errors = []

        for p in products:
            try:
                if not p.get("nome"):
                    continue
                produto = Produto(
                    nome=p["nome"],
                    tipo=p.get("tipo") or "oleo",
                    categoria=p.get("categoria"),
                    unidade=p.get("unidade"),
                    concentracao=p.get("concentracao"),
                    fabricante=p.get("fabricante"),
                    codigo_barras=p.get("codigo_barras"),
                    descricao=p.get("descricao"),
                    ativo=True,
                )
                db.session.add(produto)
                imported += 1
            except Exception as e:
                errors.append(f"Erro ao importar '{p.get('nome', '?')}': {e}")
                logger.warning(f"Falha ao importar produto: {e}")

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return {
                "success": False,
                "imported_count": 0,
                "errors": [str(e)],
            }

        # Atualiza log de importação mais recente do usuário
        if imported > 0:
            latest_log = (
                CatalogoImportLog.query
                .filter_by(user_id=user_id)
                .order_by(CatalogoImportLog.created_at.desc())
                .first()
            )
            if latest_log:
                latest_log.imported_count = imported
                db.session.commit()

        return {
            "success": True,
            "imported_count": imported,
            "errors": errors,
        }


# Instância global
extraction_service = CatalogoExtractionService()


def _infer_tipo(categoria: str) -> str:
    """Infere o campo 'tipo' do Produto a partir da categoria extraída."""
    cat = categoria.lower().strip()
    mapping = {
        "óleo": "oleo",
        "oleo": "oleo",
        "flor": "flor",
        "pomada": "pomada",
        "gummy": "gummy",
        "pet": "pet",
        "vaporizador": "vaporizador",
        "medicamento": "medicamento",
        "suplemento": "suplemento",
        "equipamento": "equipamento",
        "insumo": "insumo",
    }
    return mapping.get(cat, "oleo")
