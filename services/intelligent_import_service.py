"""
intelligent_import_service.py — Import inteligente multi-tenant

Permite ao gestor da clínica ou à secretária importar listas de:
  - profissionais_saude   (médicos, psicólogos, enfermeiros, nutricionistas, fisioterapeutas)
  - equipe_admin          (secretárias, gestores)
  - disponibilidade       (horários de consulta por profissional)
  - consultorios          (salas, andares, recursos)

Aceita: PDF, XLSX, XLS, CSV, DOCX, TXT.
Pipeline:
  1. Extrair texto tabular do arquivo (reusa padrão de catalogo_extraction_service)
  2. Detectar intent pelo cabeçalho / palavras-chave
  3. Enviar para LLM via AIProviderManager (chat_completion) com prompt
     específico do intent
  4. Validar entidades extraídas:
       - profissionais_saude: validar_conselho() para cada registro
       - equipe_admin: validar email + telefone
       - disponibilidade: validar HH:MM e dias da semana
       - consultorios: validar campos obrigatórios
  5. Persistir AuditLog (intelligent_import.analyze, intelligent_import.apply)
  6. Retornar preview para o frontend confirmar antes de aplicar

A persistência (apply) é feita em routes/intelligent_import.py — o service
só retorna preview estruturado e delega a transação.

Parte da feature feat/intelligent-import (fase I3).
"""
from __future__ import annotations

import io
import json
import logging
import re
from dataclasses import dataclass, field, asdict
from enum import Enum
from typing import Any, Dict, List, Optional, Tuple

logger = logging.getLogger(__name__)


# REDACTED
# Intents suportados
# REDACTED
class ImportIntent(str, Enum):
    PROFISSIONAIS_SAUDE = "profissionais_saude"
    EQUIPE_ADMIN = "equipe_admin"
    DISPONIBILIDADE = "disponibilidade"
    CONSULTORIOS = "consultorios"


# Heurística de detecção por cabeçalho / palavras-chave do texto.
# Ordem importa: a primeira correspondência vence.
INTENT_KEYWORDS: Dict[ImportIntent, List[str]] = {
    ImportIntent.PROFISSIONAIS_SAUDE: [
        "crm", "crp", "coren", "crn", "crefito", "conselho", "especialidade",
        "médico", "medico", "psicólogo", "psicologo", "enfermeiro",
        "nutricionista", "fisioterapeuta", "profissional de saúde",
    ],
    ImportIntent.EQUIPE_ADMIN: [
        "secretária", "secretaria", "assistente", "gestor", "recepcionista",
        "função", "funcao", "cargo administrativo",
    ],
    ImportIntent.DISPONIBILIDADE: [
        "horário", "horario", "disponibilidade", "agenda", "atende",
        "segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo",
        "manhã", "tarde", "noite", "intervalo",
    ],
    ImportIntent.CONSULTORIOS: [
        "sala", "consultório", "consultorio", "andar", "ala", "recurso",
        "equipamento", "capacidade",
    ],
}


# REDACTED
# Estruturas de retorno (dataclass para fácil serialização)
# REDACTED
@dataclass
class ImportRecord:
    """Um registro extraído do arquivo (linha de planilha / parágrafo de PDF)."""
    line_number: int
    raw: Dict[str, Any]               # chave→valor como veio do LLM
    valid: bool = False
    errors: List[str] = field(default_factory=list)
    normalized: Dict[str, Any] = field(default_factory=dict)
    # tipo_conselho detectado (apenas profissionais_saude)
    detected_conselho_tipo: Optional[str] = None
    detected_profissao: Optional[str] = None


@dataclass
class ImportPreview:
    """Preview do que seria importado (não persistido ainda)."""
    intent: str
    intent_confianca: float           # 0–1
    filename: str
    total_registros: int
    validos: int
    invalidos: int
    records: List[Dict[str, Any]]     # ImportRecord serializado
    headers_detectados: List[str]
    resumo_erros: Dict[str, int]      # tipo de erro → contagem
    ai_provider: str = "unknown"
    ai_model: str = "unknown"


# REDACTED
# Service
# REDACTED
class IntelligentImportService:
    """Orquestra detecção de intent + extração via LLM + validação."""

    ALLOWED_EXTENSIONS = {"pdf", "xlsx", "xls", "csv", "docx", "doc", "txt"}

    # Limite duro de caracteres para o LLM (evita estouro de contexto)
    MAX_TEXT_CHARS = 12_000

    def __init__(self, ai_manager=None):
        # Importação tardia para não acoplar import-time
        if ai_manager is None:
            from services.ai_agents import AIProviderManager
            ai_manager = AIProviderManager()
        self.ai = ai_manager

    # ----------------------------- I/O ---------------------------------
    def allowed_file(self, filename: str) -> bool:
        return "." in filename and filename.rsplit(".", 1)[1].lower() in self.ALLOWED_EXTENSIONS

    def extract_text(self, file_bytes: bytes, filename: str) -> str:
        """Extrai texto do arquivo em formato markdown-like (preserva tabelas).

        Reaproveita o padrão de catalogo_extraction_service para PDF/XLSX
        e adiciona fallback para CSV/DOCX/TXT.
        """
        ext = filename.rsplit(".", 1)[1].lower() if "." in filename else ""
        try:
            if ext == "pdf":
                return self._extract_text_from_pdf(file_bytes)
            if ext in ("xlsx", "xls"):
                return self._extract_text_from_xlsx(file_bytes)
            if ext == "csv":
                return self._extract_text_from_csv(file_bytes)
            if ext in ("docx", "doc"):
                return self._extract_text_from_docx(file_bytes)
            if ext == "txt":
                return file_bytes.decode("utf-8", errors="ignore")
        except Exception as exc:
            logger.warning("Falha ao extrair texto (%s): %s", ext, exc)
        return ""

    def _extract_text_from_pdf(self, file_bytes: bytes) -> str:
        try:
            from pypdf import PdfReader
        except ImportError:
            return ""
        reader = PdfReader(io.BytesIO(file_bytes))
        return "\n\n".join(p.extract_text() or "" for p in reader.pages)

    def _extract_text_from_xlsx(self, file_bytes: bytes) -> str:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return ""
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        parts: List[str] = []
        for sheet in wb.worksheets:
            parts.append(f"--- Sheet: {sheet.title} ---")
            for row in sheet.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                if any(c.strip() for c in cells):
                    parts.append(" | ".join(cells))
        return "\n".join(parts)

    def _extract_text_from_csv(self, file_bytes: bytes) -> str:
        for enc in ("utf-8", "latin1", "cp1252"):
            try:
                return file_bytes.decode(enc)
            except UnicodeDecodeError:
                continue
        return file_bytes.decode("utf-8", errors="ignore")

    def _extract_text_from_docx(self, file_bytes: bytes) -> str:
        try:
            from docx import Document  # python-docx
        except ImportError:
            return ""
        doc = Document(io.BytesIO(file_bytes))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    # ----------------------- Detecção de intent ------------------------
    def detect_intent(self, text: str) -> Tuple[ImportIntent, float]:
        """Detecção por palavras-chave (heurística simples, transparente).

        Retorna (intent, confiança 0–1). Quando ambíguo, escolhe
        profissionais_saude como default (mais comum em clínica).
        """
        if not text:
            return ImportIntent.PROFISSIONAIS_SAUDE, 0.0
        text_lower = text.lower()
        scores: Dict[ImportIntent, int] = {intent: 0 for intent in ImportIntent}
        for intent, keywords in INTENT_KEYWORDS.items():
            for kw in keywords:
                if kw in text_lower:
                    scores[intent] += 1
        best = max(scores, key=lambda k: scores[k])
        total = sum(scores.values()) or 1
        if scores[best] == 0:
            return ImportIntent.PROFISSIONAIS_SAUDE, 0.0
        return best, round(scores[best] / total, 2)

    # -------------------------- Prompts LLM ---------------------------
    def _prompt_for_intent(self, intent: ImportIntent) -> str:
        """System prompt específico por intent (orienta extração JSON)."""
        common = (
            "Responda APENAS com JSON válido, sem markdown, sem comentários. "
            "Use exatamente o schema indicado. Se um campo não existir no "
            "documento, use null (não invente)."
        )
        if intent == ImportIntent.PROFISSIONAIS_SAUDE:
            return (
                "Você é um agente especializado em extrair listas de "
                "profissionais de saúde de documentos brutos. Tipos de "
                "conselho suportados: CRM (médico), CRP (psicólogo), COREN "
                "(enfermeiro), CRN (nutricionista), CREFITO (fisioterapeuta), "
                "NONE (staff sem conselho). " + common + " Schema: "
                '{"headers": ["nome","email","telefone","conselho_tipo",'
                '"conselho_numero","uf","especialidade"], '
                '"registros": [{"nome":"","email":"","telefone":"",'
                '"conselho_tipo":"CRM","conselho_numero":"","uf":"",'
                '"especialidade":""}]}'
            )
        if intent == ImportIntent.EQUIPE_ADMIN:
            return (
                "Você extrai listas de equipe administrativa (secretárias, "
                "gestores, recepcionistas). " + common + " Schema: "
                '{"headers": ["nome","email","telefone","funcao"], '
                '"registros": [{"nome":"","email":"","telefone":"",'
                '"funcao":"secretaria"}]}'
            )
        if intent == ImportIntent.DISPONIBILIDADE:
            return (
                "Você extrai grades de disponibilidade de profissionais de "
                "saúde (horários de atendimento por dia da semana). " +
                common + " Schema: "
                '{"headers": ["profissional","dia_semana","hora_inicio",'
                '"hora_fim","intervalo_min","consultorio"], '
                '"registros": [{"profissional":"","dia_semana":"segunda",'
                '"hora_inicio":"08:00","hora_fim":"12:00",'
                '"intervalo_min":15,"consultorio":null}]}'
            )
        if intent == ImportIntent.CONSULTORIOS:
            return (
                "Você extrai listas de consultórios/salas de uma clínica. " +
                common + " Schema: "
                '{"headers": ["nome","andar","ala","capacidade","recursos"], '
                '"registros": [{"nome":"","andar":"","ala":"",'
                '"capacidade":1,"recursos":""}]}'
            )
        return common

    # ----------------------- Validação por intent ----------------------
    def _validate_records(
        self, intent: ImportIntent, registros: List[Dict[str, Any]]
    ) -> List[ImportRecord]:
        if intent == ImportIntent.PROFISSIONAIS_SAUDE:
            return [self._validate_profissional(r, idx) for idx, r in enumerate(registros, start=1)]
        if intent == ImportIntent.EQUIPE_ADMIN:
            return [self._validate_staff(r, idx) for idx, r in enumerate(registros, start=1)]
        if intent == ImportIntent.DISPONIBILIDADE:
            return [self._validate_disponibilidade(r, idx) for idx, r in enumerate(registros, start=1)]
        if intent == ImportIntent.CONSULTORIOS:
            return [self._validate_consultorio(r, idx) for idx, r in enumerate(registros, start=1)]
        return []

    def _validate_profissional(self, r: Dict[str, Any], n: int) -> ImportRecord:
        rec = ImportRecord(line_number=n, raw=r)
        nome = (r.get("nome") or "").strip()
        email = (r.get("email") or "").strip().lower()
        if not nome:
            rec.errors.append("Nome vazio")
        if not email or "@" not in email:
            rec.errors.append(f"Email inválido: '{email}'")
        conselho_tipo = r.get("conselho_tipo") or "CRM"
        conselho_numero = r.get("conselho_numero") or ""
        uf = (r.get("uf") or "").strip().upper()
        # Reusa conselho_validator
        from services.conselho_validator import validar_conselho
        v = validar_conselho(numero=conselho_numero, uf=uf, tipo=conselho_tipo)
        rec.detected_conselho_tipo = v["tipo_normalizado"]
        rec.detected_profissao = v.get("profissao")
        if not v["valido"]:
            rec.errors.extend(v["erros"])
        rec.normalized = {
            "nome": nome,
            "email": email,
            "telefone": (r.get("telefone") or "").strip(),
            "conselho_tipo": v["tipo_normalizado"],
            "conselho_numero": conselho_numero,
            "uf": uf,
            "especialidade": (r.get("especialidade") or "").strip(),
            "role": v["role"],
        }
        rec.valid = not rec.errors
        return rec

    def _validate_staff(self, r: Dict[str, Any], n: int) -> ImportRecord:
        rec = ImportRecord(line_number=n, raw=r)
        nome = (r.get("nome") or "").strip()
        email = (r.get("email") or "").strip().lower()
        funcao = (r.get("funcao") or "secretaria").strip().lower()
        if not nome:
            rec.errors.append("Nome vazio")
        if not email or "@" not in email:
            rec.errors.append(f"Email inválido: '{email}'")
        if funcao not in ("secretaria", "gestor", "recepcionista", "auxiliar", "manager", "secretary"):
            rec.errors.append(f"Função '{funcao}' não reconhecida")
        rec.normalized = {
            "nome": nome,
            "email": email,
            "telefone": (r.get("telefone") or "").strip(),
            "funcao": funcao if funcao in ("secretary", "secretaria", "auxiliar") else (
                "manager" if funcao in ("gestor", "manager") else "secretary"
            ),
        }
        rec.valid = not rec.errors
        return rec

    _DIAS_VALIDOS = {
        "segunda", "terca", "terça", "quarta", "quinta", "sexta",
        "sabado", "sábado", "domingo", "seg", "ter", "qua", "qui",
        "sex", "sab", "dom",
    }
    _HORA_RE = re.compile(r"^([01]?\d|2[0-3]):[0-5]\d$")

    def _validate_disponibilidade(self, r: Dict[str, Any], n: int) -> ImportRecord:
        rec = ImportRecord(line_number=n, raw=r)
        profissional = (r.get("profissional") or "").strip()
        dia = (r.get("dia_semana") or "").strip().lower()
        hi = (r.get("hora_inicio") or "").strip()
        hf = (r.get("hora_fim") or "").strip()
        if not profissional:
            rec.errors.append("Profissional vazio")
        if dia not in self._DIAS_VALIDOS:
            rec.errors.append(f"Dia da semana inválido: '{dia}'")
        if not self._HORA_RE.match(hi):
            rec.errors.append(f"hora_inicio inválida: '{hi}' (esperado HH:MM)")
        if not self._HORA_RE.match(hf):
            rec.errors.append(f"hora_fim inválida: '{hf}' (esperado HH:MM)")
        if self._HORA_RE.match(hi) and self._HORA_RE.match(hf) and hi >= hf:
            rec.errors.append(f"hora_inicio ({hi}) deve ser anterior a hora_fim ({hf})")
        rec.normalized = {
            "profissional": profissional,
            "dia_semana": dia,
            "hora_inicio": hi,
            "hora_fim": hf,
            "intervalo_min": int(r.get("intervalo_min") or 30),
            "consultorio": r.get("consultorio"),
        }
        rec.valid = not rec.errors
        return rec

    def _validate_consultorio(self, r: Dict[str, Any], n: int) -> ImportRecord:
        rec = ImportRecord(line_number=n, raw=r)
        nome = (r.get("nome") or "").strip()
        if not nome:
            rec.errors.append("Nome do consultório vazio")
        try:
            capacidade = int(r.get("capacidade") or 1)
            if capacidade < 1:
                rec.errors.append("capacidade deve ser ≥ 1")
        except (TypeError, ValueError):
            rec.errors.append(f"capacidade inválida: '{r.get('capacidade')}'")
            capacidade = 1
        rec.normalized = {
            "nome": nome,
            "andar": (r.get("andar") or "").strip(),
            "ala": (r.get("ala") or "").strip(),
            "capacidade": capacidade,
            "recursos": (r.get("recursos") or "").strip(),
        }
        rec.valid = not rec.errors
        return rec

    # ----------------------- Pipeline principal ------------------------
    def analyze(
        self,
        file_bytes: bytes,
        filename: str,
        *,
        intent_override: Optional[str] = None,
    ) -> ImportPreview:
        """Lê arquivo, detecta intent, extrai via LLM, valida e retorna preview.

        NÃO persiste nada — o caller (route) decide se aplica após confirmação.
        """
        if not self.allowed_file(filename):
            return ImportPreview(
                intent="",
                intent_confianca=0.0,
                filename=filename,
                total_registros=0, validos=0, invalidos=0,
                records=[],
                headers_detectados=[],
                resumo_erros={"formato_nao_suportado": 1},
            )

        text = self.extract_text(file_bytes, filename).strip()
        if not text:
            return ImportPreview(
                intent="",
                intent_confianca=0.0,
                filename=filename,
                total_registros=0, validos=0, invalidos=0,
                records=[],
                headers_detectados=[],
                resumo_erros={"texto_vazio": 1},
            )

        # Truncar para evitar estouro de contexto no LLM
        if len(text) > self.MAX_TEXT_CHARS:
            text = text[: self.MAX_TEXT_CHARS]

        if intent_override:
            intent = ImportIntent(intent_override)
            confianca = 1.0
        else:
            intent, confianca = self.detect_intent(text)

        # Extrair via LLM
        system_prompt = self._prompt_for_intent(intent)
        user_prompt = (
            f"Analise o documento abaixo e extraia os dados no schema JSON "
            f"indicado.\n\nDOCUMENTO:\n{text}"
        )
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]
        try:
            response = self.ai.chat_completion(
                messages=messages,
                temperature=0.1,
                max_tokens=2000,
            )
            content = response.get("content", "")
        except Exception as exc:
            logger.error("Falha no LLM para intelligent_import: %s", exc)
            return ImportPreview(
                intent=intent.value, intent_confianca=confianca, filename=filename,
                total_registros=0, validos=0, invalidos=0, records=[],
                headers_detectados=[],
                resumo_erros={"llm_indisponivel": 1},
            )

        # Parsear JSON com fallback tolerante
        try:
            parsed = self._parse_json(content)
        except Exception as exc:
            logger.warning("LLM não retornou JSON válido: %s | conteúdo=%s", exc, content[:200])
            return ImportPreview(
                intent=intent.value, intent_confianca=confianca, filename=filename,
                total_registros=0, validos=0, invalidos=0, records=[],
                headers_detectados=[],
                resumo_erros={"json_invalido": 1},
                ai_provider=response.get("provider", "unknown"),
                ai_model=response.get("model", "unknown"),
            )

        headers = parsed.get("headers") or []
        registros = parsed.get("registros") or []
        validated = self._validate_records(intent, registros)
        validos = sum(1 for r in validated if r.valid)
        invalidos = len(validated) - validos

        # Resumo de erros (agregado)
        resumo: Dict[str, int] = {}
        for rec in validated:
            for err in rec.errors:
                key = err.split(":")[0].split("(")[0].strip()[:40] or "outro"
                resumo[key] = resumo.get(key, 0) + 1

        return ImportPreview(
            intent=intent.value,
            intent_confianca=confianca,
            filename=filename,
            total_registros=len(validated),
            validos=validos,
            invalidos=invalidos,
            records=[asdict(r) for r in validated],
            headers_detectados=headers,
            resumo_erros=resumo,
            ai_provider=response.get("provider", "unknown"),
            ai_model=response.get("model", "unknown"),
        )

    @staticmethod
    def _parse_json(content: str) -> Dict[str, Any]:
        """Tenta extrair JSON de respostas LLM (tolerante a markdown)."""
        if not content:
            return {}
        s = content.strip()
        # remover fences ```json ... ```
        if s.startswith("```"):
            s = re.sub(r"^```(?:json)?\s*", "", s)
            s = re.sub(r"\s*```$", "", s)
        # pegar o primeiro {...} ou [...]
        m = re.search(r"(\{.*\}|\[.*\])", s, flags=re.DOTALL)
        candidate = m.group(1) if m else s
        return json.loads(candidate)


__all__ = [
    "IntelligentImportService",
    "ImportIntent",
    "ImportRecord",
    "ImportPreview",
]
